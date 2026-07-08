"""
Le planilhas Excel de manutencao aeronautica e gera data.json para o dashboard.

Replica fielmente a logica de parsing do dashboard.html (parseWorkbook em JS).
"""

import json
import os
import re
import unicodedata
from datetime import datetime, date
from pathlib import Path

import xlrd  # leitura de .xls
from openpyxl import load_workbook  # leitura de .xlsx

# ── CONFIG ──────────────────────────────────────────────────────────────────

# Pasta base onde estao as subpastas de cada aeronave. Pode ser sobrescrita
# pela variavel de ambiente ONEDRIVE_BASE (usado no workflow do GitHub Actions,
# que sincroniza o OneDrive para um diretorio do runner).
DEFAULT_BASE = r"C:\Users\ysado\OneDrive"
BASE_DIR = Path(os.environ.get("ONEDRIVE_BASE", DEFAULT_BASE))

# Cada aeronave tem uma subpasta; dentro dela pegamos o arquivo .xls/.xlsx
# mais recente (por data de modificacao), pois o nome do arquivo muda a cada
# atualizacao da planilha (ex: "... - 14-06-2026.xls").
AIRCRAFT_FOLDERS = [
    "Planilhas 2026 - PP-AGN",
    "Planilhas 2026 - PP-VEL",
    "Planilhas 2026 - PS-FLC",
    "Planilhas 2026 - PS-NFA",
    # Novas aeronaves
    "Planilhas 2026",                                        # King Air B300 - PS-KNG
    "Planilhas 2026 - G700 - N444R",                        # Gulfstream G700
    "Planilhas 2026 - PS-JAJ",                              # Praetor 600
    "Planilhas 2026 - PS-STP",                              # Global 5000
    "Gulfstream G450-N918LL/Planilhas 2026 - N918LL",       # Gulfstream G450
]


def find_latest_spreadsheet(folder: Path):
    candidates = [
        f for f in folder.glob("*")
        if f.suffix.lower() in (".xls", ".xlsx") and not f.name.startswith("~$")
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda f: f.stat().st_mtime)

SITE_DIR = Path(__file__).resolve().parent / "site"
DATA_DIR  = SITE_DIR / "data"       # per-aircraft JSON files served via Worker
OUTPUT_PATH    = SITE_DIR / "data.json"   # kept for legacy / local testing
OUTPUT_JS_PATH = SITE_DIR / "data.js"     # kept for legacy / local testing

TARGET_SHEETS = ["Manutenção", "Manutenao", "Componentes", "DIR", "DIR MOTOR", "DIR APU", "Diário de Bordo"]

# Matches ANAC registrations (PP-AGN) and FAA registrations (N444R, N918LL)
ACFT_NAME_RE = re.compile(r"\b([A-Z]{2}-[A-Z0-9]{3}|N\d{1,5}[A-Z]{0,2})\b", re.IGNORECASE)


# ── HELPERS ──────────────────────────────────────────────────────────────────

def normalize(v):
    s = "" if v is None else str(v)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.strip().upper()


def is_alert(v):
    n = normalize(v)
    return n in ("ATENCAO", "ATTENTION")


def excel_date_to_py(serial):
    if serial is None or not isinstance(serial, (int, float)):
        return None
    try:
        from datetime import timedelta
        return date(1899, 12, 30) + timedelta(days=serial)
    except Exception:
        return None


def to_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        return excel_date_to_py(val)
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
        return None
    return None


def date_diff_days(val):
    d = to_date(val)
    if d is None:
        return None
    today = date.today()
    return (d - today).days


def extract_acft_name(folder_name: str, file_name: str):
    # Try folder name first (more reliable: user chose folder names with registration)
    for name in (folder_name, file_name):
        m = ACFT_NAME_RE.search(name)
        if m:
            return m.group(1).upper()
    base = Path(file_name).stem
    b = re.search(r"__([A-Z0-9]{5})__", base, re.IGNORECASE)
    if b:
        c = b.group(1).upper()
        return c[:2] + "-" + c[2:]
    return re.split(r"[_\-\s]+", base)[0].upper()


def num(v):
    try:
        if v is None or v in ("", "-", "--"):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def fmt_num(v):
    """Valor pronto para exibicao: mantem string original, ou formata numero
    cortando ruido de ponto flutuante (ex: 2204.8000000000006 -> '2204.8')."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return None if s in ("", "-", "--") else s
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v)
    r = round(f, 1)
    return str(int(r)) if r == int(r) else f"{r:.1f}"


# ── LEITURA DE WORKBOOK (abstrai .xls vs .xlsx) ─────────────────────────────

class Sheet:
    def __init__(self, name, rows):
        self.name = name
        self.rows = rows  # list of list, header:1 style (0-indexed rows/cols)


def read_workbook(path):
    """Retorna lista de Sheet com linhas em formato lista-de-listas (defval=None)."""
    ext = Path(path).suffix.lower()
    sheets = []
    if ext == ".xls":
        wb = xlrd.open_workbook(path)
        for sn in wb.sheet_names():
            ws = wb.sheet_by_name(sn)
            rows = []
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    val = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            tup = xlrd.xldate_as_tuple(val, wb.datemode)
                            val = datetime(*tup)
                        except Exception:
                            pass
                    elif val == "":
                        val = None
                    row.append(val)
                rows.append(row)
            sheets.append(Sheet(sn, rows))
    else:
        wb = load_workbook(path, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = []
            for row_cells in ws.iter_rows(values_only=True):
                rows.append(list(row_cells))
            sheets.append(Sheet(sn, rows))
    return sheets


def get(row, idx):
    if idx < len(row):
        return row[idx]
    return None


# ── PARSER (espelha parseWorkbook do JS) ────────────────────────────────────

def parse_workbook(sheets, acft_name):
    tasks = []
    info = {}

    for sheet in sheets:
        sn = sheet.name.strip()
        if not any(sn == t or sn.startswith(t) for t in TARGET_SHEETS):
            continue

        rows = sheet.rows

        if normalize(sn).startswith("DIARIO DE BORDO"):
            fl = parse_flight_log(rows)
            if fl is not None:
                info["flightLog"] = fl
            continue

        # Modelo da aeronave: primeira célula não-vazia da linha 0
        if normalize(sn).startswith("MANUTENCAO") and "model" not in info:
            row0 = rows[0] if rows else []
            for cell in row0:
                s = str(cell).strip() if cell is not None else ""
                if s:
                    info["model"] = s
                    break

        # Cabecalho da aeronave
        if "totalHours" not in info:
            for i in range(min(15, len(rows))):
                row = rows[i] or []
                for j in range(len(row) - 1):
                    v = normalize(get(row, j))
                    if v == "HORAS TOTAIS":
                        n = num(get(row, j + 1))
                        if n is not None:
                            info["totalHours"] = n
                    if v == "POUSOS TOTAIS":
                        n = num(get(row, j + 1))
                        if n is not None:
                            info["totalLandings"] = n
                    if v == "CICLOS TOTAIS":
                        n = num(get(row, j + 1))
                        if n is not None:
                            info["totalCycles"] = n
                    if "registration" not in info:
                        cell_str = str(get(row, j) or "")
                        rm = re.match(r"^([A-Z]{2}-[A-Z0-9]{3})$", cell_str, re.IGNORECASE)
                        if rm:
                            info["registration"] = rm.group(1).upper()

        is_dir = sn.startswith("DIR")

        if is_dir:
            hdr = -1
            for i, row in enumerate(rows):
                rs = "|".join(normalize(v) for v in (row or []))
                if ("AIRWORTHINESS" in rs) or ("DIRETRIZ" in rs) or ("AD/DA" in rs) or ("AD DA" in rs):
                    hdr = i
                    break
            if hdr < 0:
                continue
            for i in range(hdr + 1, len(rows)):
                row = rows[i] or []
                alert_h = is_alert(get(row, 14))
                alert_d = is_alert(get(row, 15))
                if not alert_h and not alert_d:
                    continue
                ad = str(get(row, 1) or "").strip()
                sb = str(get(row, 2) or "").strip()
                desc = str(get(row, 3) or "").strip()
                if not ad and not sb and not desc:
                    continue

                alert_types = []
                due_hours_str = None
                due_days_str = None

                if alert_h and get(row, 12) is not None:
                    h = num(get(row, 12))
                    if h is not None:
                        due_hours_str = f"{h:.1f}h"
                        alert_types.append("hours")

                if alert_d:
                    days = date_diff_days(get(row, 13))
                    if days is not None:
                        due_days_str = f"{days} dias"
                        alert_types.append("days")
                    elif get(row, 13) is not None:
                        n = num(get(row, 13))
                        if n is not None:
                            due_days_str = f"{round(n)} dias"
                            alert_types.append("days")

                tasks.append({
                    "id": str(get(row, 0) or "").strip(),
                    "task": ad or sb or "",
                    "description": desc or ad or "AD",
                    "pn": sb or "",
                    "sheet": sn,
                    "dueHoursStr": due_hours_str,
                    "dueDaysStr": due_days_str,
                    "dueCyclesStr": None,
                    "alertTypes": alert_types,
                })
        else:
            hdr = -1
            for i, row in enumerate(rows):
                rs = "|".join(normalize(v) for v in (row or []))
                if ("TASK" in rs) or ("ID" in rs and ("INSPECTIONS" in rs or "NOMENCLATURA" in rs)):
                    hdr = i
                    break
            if hdr < 0:
                continue

            alert_col = 19 if sn == "Componentes" else 18

            for i in range(hdr + 1, len(rows)):
                row = rows[i] or []
                if not is_alert(get(row, alert_col)):
                    continue
                task_id = str(get(row, 1) or "").strip()
                desc = str(get(row, 2) or "").strip()
                if not task_id and not desc:
                    continue

                alert_types = []
                due_hours_str = None
                due_days_str = None
                due_cycles_str = None

                # Horas: col[14] direto (saldo calculado pelo Excel)
                saldo_h = get(row, 14)
                if saldo_h is not None and saldo_h != "-" and saldo_h != "":
                    h = num(saldo_h)
                    if h is not None:
                        due_hours_str = f"{h:.1f}h"
                        alert_types.append("hours")

                # Dias: recalcular com col[13] (Date)
                next_date = get(row, 13)
                if next_date is not None and next_date != "-":
                    days = date_diff_days(next_date)
                    if days is not None:
                        due_days_str = f"{days} dias"
                        alert_types.append("days")
                    else:
                        sd = get(row, 16)
                        if sd is not None and sd != "-":
                            n = num(sd)
                            if n is not None:
                                due_days_str = f"{round(n)} dias"
                                alert_types.append("days")
                else:
                    sd = get(row, 16)
                    if sd is not None and sd != "-" and sd != "":
                        n = num(sd)
                        if n is not None:
                            due_days_str = f"{round(n)} dias"
                            alert_types.append("days")

                # Ciclos: col[15]
                saldo_c = get(row, 15)
                if saldo_c is not None and saldo_c != "-" and saldo_c != "":
                    c = num(saldo_c)
                    if c is not None:
                        due_cycles_str = f"{round(c)} ciclos"
                        alert_types.append("cycles")

                if not alert_types:
                    alert_types.append("days")

                tasks.append({
                    "id": str(get(row, 0) or "").strip(),
                    "task": task_id,
                    "description": desc,
                    "pn": str(get(row, 3) or "").strip(),
                    "sheet": sn,
                    "dueHoursStr": due_hours_str,
                    "dueDaysStr": due_days_str,
                    "dueCyclesStr": due_cycles_str,
                    "alertTypes": alert_types,
                })

    info.setdefault("model", None)
    info.setdefault("totalHours", None)
    info.setdefault("totalLandings", None)
    info.setdefault("totalCycles", None)
    info.setdefault("flightLog", None)

    return {"tasks": tasks, "info": info, "name": acft_name}


# ── DIÁRIO DE BORDO (Flight Log) ────────────────────────────────────────────
# A aba "Diário de Bordo" registra o historico de voo etapa-a-etapa (celula,
# motores, pousos, ciclos e, em algumas aeronaves, APU) desde a entrada da
# aeronave na frota. O layout de colunas varia por aeronave (motor unico vs.
# LH/RH, APU com 0/2/4 campos) e ate entre "livros" diferentes da mesma
# aeronave, entao as colunas sao localizadas pelo ROTULO do cabecalho (nao por
# indice fixo).

FLIGHT_LOG_TITLE_RE = re.compile(
    r"(DIARIO DE BORDO|FLIGHT LOG(?:BOOK)?)\s+N[O0º]\.?\s*\d+",
    re.IGNORECASE,
)

MESES_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

APU_QUAD_KEYS = {
    ("AH", "ANALOG"): "ahAnalog",
    ("AH", "DIGITAL"): "ahDigital",
    ("AC", "ANALOG"): "acAnalog",
    ("AC", "DIGITAL"): "acDigital",
}


def _fwd_fill_labels(row):
    """Propaga o rotulo de cada celula mesclada (nao vazia) para as celulas
    vazias seguintes, simulando o merge visual do Excel."""
    out = []
    current = ""
    for v in (row or []):
        n = normalize(v)
        if n:
            current = n
        out.append(current)
    return out


def _resolve_flight_log_columns(title_row, group_row, sub_row):
    """Localiza por rotulo as colunas relevantes de uma secao do Diario de
    Bordo. Retorna (cols, apu_cols); chaves ausentes em `cols` = coluna nao
    existe nesta aeronave/secao."""
    width = max(len(group_row or []), len(sub_row or []))
    groups = _fwd_fill_labels(group_row)  # p/ blocos mesclados (celula/motor)
    raw = [normalize(v) for v in (group_row or [])]  # p/ celulas isoladas (data/apu, nao mescladas)
    title_groups = _fwd_fill_labels(title_row)
    cols = {}
    apu_cols = []  # (idx, 'AH'|'AC', 'ANALOG'|'DIGITAL'|None)

    for i in range(width):
        g = groups[i] if i < len(groups) else ""
        r = raw[i] if i < len(raw) else ""
        s = normalize(get(sub_row, i))

        if r in ("DATA", "DATE"):
            cols.setdefault("date", i)
            continue

        if s in ("FOLHA", "PAGE"):
            cols["folha"] = i
            continue

        if "APU" in r:
            if "AH" in r or "FH" in r or "HORAS" in r:
                kind = "AH"
            elif "AC" in r or "CICLO" in r or "CYCLE" in r:
                kind = "AC"
            else:
                continue
            marker = title_groups[i] if i < len(title_groups) else ""
            sub = "ANALOG" if "ANALOG" in marker else ("DIGITAL" if "DIGIT" in marker else None)
            apu_cols.append((i, kind, sub))
            continue

        if "CELULA" in g or "AIRFRAME" in g:
            if "TEMPO" in g or "HOURS" in g:
                if s == "HS/MI": cols["celulaEtapa"] = i
                elif s == "TOTAL": cols["celulaTotal"] = i
            elif "POUSO" in g or "CYCLE" in g:
                if s in ("ETAPA", "CYCLES"): cols["pousoEtapa"] = i
                elif s in ("TOTAIS", "TOTAL"): cols["pousoTotal"] = i
            continue

        if "MOTOR" in g or "ENGINE" in g:
            if "TEMPO" in g or "HOURS" in g:
                if s == "HS/MI": cols["motorEtapa"] = i
                elif s == "TOTAL": cols["motorTotal"] = i
                elif s == "LH": cols["lh"] = i
                elif s == "RH": cols["rh"] = i
            elif "CICLO" in g or "CYCLE" in g:
                if "LH" in g:
                    if s == "CT": cols["lhCT"] = i
                    elif s == "PT": cols["lhPT"] = i
                    elif s in ("IMP", "MP"): cols["lhIMP"] = i
                elif "RH" in g:
                    if s == "CT": cols["rhCT"] = i
                    elif s == "PT": cols["rhPT"] = i
                    elif s in ("IMP", "MP"): cols["rhIMP"] = i
                else:
                    if s in ("ETAPA", "CYCLES"): cols["cicloEtapa"] = i
                    elif s in ("TOTAIS", "TOTAL"): cols["cicloTotal"] = i
            continue

    # A coluna DATA nem sempre reimprime o rotulo em livros repetidos do
    # mesmo Diario de Bordo (o valor continua la, so o texto do cabecalho
    # some). Ela fica de forma consistente logo apos o ultimo bloco de
    # motor/ciclos, entao usamos a posicao como referencia quando o rotulo
    # nao for encontrado.
    if "date" not in cols:
        motor_end = max(
            (cols[k] for k in ("cicloTotal", "rhIMP", "lhIMP", "motorTotal", "celulaTotal") if k in cols),
            default=None,
        )
        if motor_end is not None:
            cols["date"] = motor_end + 1

    return cols, apu_cols


def parse_flight_log(rows):
    title_idxs = [
        i for i, row in enumerate(rows)
        if row and FLIGHT_LOG_TITLE_RE.search(normalize(get(row, 0)) or "")
    ]
    if not title_idxs:
        return None

    legs = []
    twin = False
    apu_mode = "none"
    apu_rank = {"none": 0, "single": 1, "quad": 2}

    for si, start in enumerate(title_idxs):
        end = title_idxs[si + 1] if si + 1 < len(title_idxs) else len(rows)
        title_text = str(get(rows[start], 0) or "").strip()

        sub_idx = None
        for i in range(start + 1, min(start + 6, end)):
            if normalize(get(rows[i], 0)) in ("FOLHA", "PAGE"):
                sub_idx = i
                break
        if sub_idx is None or sub_idx - 1 < start:
            continue

        group_row = rows[sub_idx - 1] or []
        sub_row = rows[sub_idx] or []
        title_marker_row = rows[start] or []
        cols, apu_cols = _resolve_flight_log_columns(title_marker_row, group_row, sub_row)

        if "lh" in cols or "rh" in cols:
            twin = True
        if apu_cols:
            has_marker = any(sub for _, _, sub in apu_cols)
            new_mode = "quad" if has_marker else "single"
            if apu_rank[new_mode] > apu_rank[apu_mode]:
                apu_mode = new_mode

        folha_col = cols.get("folha", 0)
        date_col = cols.get("date")
        quad_this_section = any(sub for _, _, sub in apu_cols)

        for ri in range(sub_idx + 1, end):
            row = rows[ri] or []
            folha_val = get(row, folha_col)
            if folha_val is None or str(folha_val).strip() == "":
                continue
            celula_total_col = cols.get("celulaTotal")
            if celula_total_col is not None and get(row, celula_total_col) is None:
                continue

            d = to_date(get(row, date_col)) if date_col is not None else None

            leg = {
                "book": title_text,
                "folha": fmt_num(folha_val),
                "date": d.strftime("%d/%m/%Y") if d else None,
                "_date": d,
                "celulaEtapa": fmt_num(get(row, cols.get("celulaEtapa"))),
                "celulaTotal": fmt_num(get(row, cols.get("celulaTotal"))),
                "pousoEtapa": fmt_num(get(row, cols.get("pousoEtapa"))),
                "pousoTotal": fmt_num(get(row, cols.get("pousoTotal"))),
                "_hoursDelta": num(get(row, cols.get("celulaEtapa"))),
                "_landingsDelta": num(get(row, cols.get("pousoEtapa"))),
            }

            if "lh" in cols or "rh" in cols:
                leg["motorEtapa"] = fmt_num(get(row, cols.get("motorEtapa")))
                leg["lh"] = fmt_num(get(row, cols.get("lh")))
                leg["rh"] = fmt_num(get(row, cols.get("rh")))
                leg["lhCT"] = fmt_num(get(row, cols.get("lhCT")))
                leg["lhPT"] = fmt_num(get(row, cols.get("lhPT")))
                leg["lhIMP"] = fmt_num(get(row, cols.get("lhIMP")))
                leg["rhCT"] = fmt_num(get(row, cols.get("rhCT")))
                leg["rhPT"] = fmt_num(get(row, cols.get("rhPT")))
                leg["rhIMP"] = fmt_num(get(row, cols.get("rhIMP")))
            else:
                leg["motorEtapa"] = fmt_num(get(row, cols.get("motorEtapa")))
                leg["motorTotal"] = fmt_num(get(row, cols.get("motorTotal")))
                leg["cicloEtapa"] = fmt_num(get(row, cols.get("cicloEtapa")))
                leg["cicloTotal"] = fmt_num(get(row, cols.get("cicloTotal")))

            for idx, kind, sub in apu_cols:
                val = fmt_num(get(row, idx))
                if quad_this_section:
                    key = APU_QUAD_KEYS.get((kind, sub))
                else:
                    key = "apuAH" if kind == "AH" else "apuAC"
                if key:
                    leg[key] = val

            legs.append(leg)

    if not legs:
        return None

    return build_flight_log(legs, twin, apu_mode)


def _export_leg(leg):
    return {k: v for k, v in leg.items() if not k.startswith("_")}


def build_flight_log(legs, twin, apu_mode):
    dated = [l for l in legs if l.get("_date")]
    if not dated:
        return None
    dated.sort(key=lambda l: l["_date"])  # ascendente (estavel -> ordem original nos empates)

    last = dated[-1]

    years_map = {}
    for l in dated:
        years_map.setdefault(l["_date"].year, {}).setdefault(l["_date"].month, []).append(l)

    years = []
    for y in sorted(years_map.keys(), reverse=True):
        months_map = years_map[y]
        months = []
        year_hours, year_landings, year_legs = 0.0, 0.0, 0
        for m in sorted(months_map.keys(), reverse=True):
            month_legs = sorted(months_map[m], key=lambda l: l["_date"], reverse=True)
            mh = sum(l["_hoursDelta"] for l in month_legs if l["_hoursDelta"] is not None)
            ml = sum(l["_landingsDelta"] for l in month_legs if l["_landingsDelta"] is not None)
            year_hours += mh
            year_landings += ml
            year_legs += len(month_legs)
            months.append({
                "month": m,
                "monthLabel": MESES_PT[m],
                "hours": fmt_num(mh),
                "landings": fmt_num(ml),
                "legCount": len(month_legs),
                "legs": [_export_leg(l) for l in month_legs],
            })
        years.append({
            "year": y,
            "hours": fmt_num(year_hours),
            "landings": fmt_num(year_landings),
            "legCount": year_legs,
            "months": months,
        })

    return {
        "totalHours": last.get("celulaTotal"),
        "totalLandings": last.get("pousoTotal"),
        "twin": twin,
        "apuMode": apu_mode,
        "years": years,
    }


# ── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    aircraft = {}

    for folder_name in AIRCRAFT_FOLDERS:
        folder = BASE_DIR / Path(folder_name)  # Path() handles "/" sub-paths cross-platform
        if not folder.exists():
            print(f"[AVISO] pasta nao encontrada, ignorando: {folder}")
            continue

        p = find_latest_spreadsheet(folder)
        if p is None:
            print(f"[AVISO] nenhuma planilha encontrada em: {folder}")
            continue

        acft_name = extract_acft_name(folder_name, p.name)
        print(f"Lendo {p.name} -> {acft_name}")
        try:
            sheets = read_workbook(str(p))
            data = parse_workbook(sheets, acft_name)
            aircraft[acft_name] = {"info": data["info"], "tasks": data["tasks"]}
            print(f"  {len(data['tasks'])} alerta(s) encontrado(s)")
        except Exception as e:
            print(f"[ERRO] falha ao processar {p.name}: {e}")

    output = {
        "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "aircraft": aircraft,
    }

    json_text = json.dumps(output, ensure_ascii=False, indent=2)

    # Legacy files (kept for local testing / fallback)
    OUTPUT_PATH.write_text(json_text, encoding="utf-8")
    OUTPUT_JS_PATH.write_text(f"window.EMBEDDED_DATA = {json_text};\n", encoding="utf-8")
    print(f"\nGerado (legado): {OUTPUT_PATH}")

    # Per-aircraft files served via Cloudflare Worker auth
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    owner_path = DATA_DIR / "owner.json"
    owner_path.write_text(json_text, encoding="utf-8")
    print(f"Gerado: {owner_path}")

    for reg, acft_data in output["aircraft"].items():
        single = {"generated_at": output["generated_at"], "aircraft": {reg: acft_data}}
        single_text = json.dumps(single, ensure_ascii=False, indent=2)
        acft_path = DATA_DIR / f"{reg}.json"
        acft_path.write_text(single_text, encoding="utf-8")
        print(f"Gerado: {acft_path}")


if __name__ == "__main__":
    main()
